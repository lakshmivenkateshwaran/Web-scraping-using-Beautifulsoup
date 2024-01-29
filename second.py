#Code to retrive the top 250 movies from IMDB Website using beautifulsoup
#Also we have used Openpyxl package

from bs4 import BeautifulSoup
import requests
import openpyxl
import time
import datetime


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie_Name', 'Year', 'Rating'])

URL = 'https://www.imdb.com/chart/top/'

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"}

page = requests.get(URL, headers=headers)

soup1 = BeautifulSoup(page.text, "html.parser")
movies = soup1.find(
    'ul', class_="ipc-metadata-list ipc-metadata-list--dividers-between sc-3a353071-0 wTPeg compact-list-view ipc-metadata-list--base").find_all('li')
for movie in movies:
    names = movie.find('div', class_="sc-14dd939d-0 fBusXE cli-children").text
    name = movie.find('div', class_="sc-14dd939d-0 fBusXE cli-children").a.text
    year = movie.find(
        'div', class_="sc-14dd939d-5 cPiUKY cli-title-metadata").span.text
    Rating = movie.find(
        'span', class_="ipc-rating-star ipc-rating-star--base ipc-rating-star--imdb ratingGroup--imdb-rating").text
    print(name, year, Rating)
    sheet.append([name, year, Rating])
excel.save('IMDB Top Rated Movies.xlsx')
# print(names)
# name = movie.find('a', class_ = "ipc-title-link-wrapper").h3.text()
# print(len(name))
# print(len(movies))
